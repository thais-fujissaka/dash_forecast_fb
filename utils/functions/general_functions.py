import streamlit as st
import pandas as pd
import numpy as np
from utils.queries import *
import mysql.connector
from workalendar.america import Brazil
import openpyxl
import os


def mysql_connection_fb():
  mysql_config = st.secrets["mysql_fb"]

  conn_fb = mysql.connector.connect(
        host=mysql_config['host'],
        port=mysql_config['port'],
        database=mysql_config['database'],
        user=mysql_config['username'],
        password=mysql_config['password']
    )    
  return conn_fb


def execute_query(query, conn):
    cursor = conn.cursor()
    cursor.execute(query)

    # Obter nomes das colunas
    column_names = [col[0] for col in cursor.description]
  
    # Obter resultados
    result = cursor.fetchall()
  
    cursor.close()
    return result, column_names


def dataframe_query(query):
  resultado, nomeColunas = execute_query(query)
  dataframe = pd.DataFrame(resultado, columns=nomeColunas)
  return dataframe


def export_to_excel(df, sheet_name, excel_filename):
  if os.path.exists(excel_filename):
    wb = openpyxl.load_workbook(excel_filename)
  else:
    wb = openpyxl.Workbook()

  if sheet_name in wb.sheetnames:
    wb.remove(wb[sheet_name])
  ws = wb.create_sheet(title=sheet_name)
  
  # Escrever os cabeçalhos
  for col_idx, column_title in enumerate(df.columns, start=1):
    ws.cell(row=1, column=col_idx, value=column_title)
  
  # Escrever os dados
  for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
    for col_idx, value in enumerate(row, start=1):
      ws.cell(row=row_idx, column=col_idx, value=value)

  wb.save(excel_filename)


def format_brazilian(num):
  try:
    num = float(num)
    return f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
  except (ValueError, TypeError):
    return num


def format_columns_brazilian(df, numeric_columns):
  for col in numeric_columns:
    if col in df.columns:
      df[col] = df[col].apply(format_brazilian)
  return df


def format_percentage(num):
  try:
    num = float(num)
    formatted_num = f"{num * 100:,.2f}"  # Multiplica por 100 e formata
    return f"{formatted_num.replace(',', 'X').replace('.', ',').replace('X', '.')}%"  # Formata como percentual
  except (ValueError, TypeError):
    return num  # Retorna o valor original em caso de erro
  

def format_columns_percentage(df, numeric_columns):
  for col in numeric_columns:
    if col in df.columns:
      df[col] = df[col].apply(format_percentage)
  return df  